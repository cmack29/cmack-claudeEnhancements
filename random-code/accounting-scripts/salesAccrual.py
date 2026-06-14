"""
Sales Accrual Calculator
========================
Reads invoices/orders from a CSV or Excel file, calculates accruals
for a target period, and exports results to a formatted Excel workbook.

Expected input columns (flexible — script auto-maps common variants):
  - invoice_id / order_id / id
  - customer / client / account
  - invoice_date / order_date / date
  - due_date / payment_date (optional)
  - amount / revenue / value / total
  - status (optional: paid / unpaid / partial)
  - description / notes (optional)

Usage:
  python sales_accrual.py --input sales_data.csv --period 2025-06
  python sales_accrual.py --input orders.xlsx --period 2025-Q2
  python sales_accrual.py --input sales_data.csv --period 2025-06 --output my_accrual.xlsx
"""

import argparse
import sys
from pathlib import Path
from datetime import datetime, date
import calendar

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Column name normalisation
# ---------------------------------------------------------------------------

COLUMN_ALIASES = {
    "invoice_id":   ["invoice_id", "order_id", "id", "ref", "reference", "invoice_no", "order_no"],
    "customer":     ["customer", "client", "account", "customer_name", "client_name", "company"],
    "invoice_date": ["invoice_date", "order_date", "date", "sale_date", "transaction_date", "raised_date"],
    "due_date":     ["due_date", "payment_date", "payment_due", "due"],
    "amount":       ["amount", "revenue", "value", "total", "net", "net_amount", "sales_amount", "price"],
    "status":       ["status", "payment_status", "state"],
    "description":  ["description", "notes", "note", "details", "item", "product"],
}


def normalise_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Rename columns to standard names based on aliases."""
    rename_map = {}
    lower_cols = {c.lower().replace(" ", "_"): c for c in df.columns}
    for std_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in lower_cols and std_name not in rename_map.values():
                rename_map[lower_cols[alias]] = std_name
                break
    return df.rename(columns=rename_map)


# ---------------------------------------------------------------------------
# Period parsing
# ---------------------------------------------------------------------------

def parse_period(period_str: str):
    """
    Parse a period string into (start_date, end_date, label).
    Supports:
      2025-06        → June 2025
      2025-Q2        → Q2 2025 (Apr–Jun)
      2025           → Full year 2025
    """
    s = period_str.strip().upper()

    if "Q" in s:
        year, q = s.split("-Q")
        year = int(year)
        q = int(q)
        if q not in (1, 2, 3, 4):
            raise ValueError(f"Quarter must be 1-4, got {q}")
        quarter_starts = {1: 1, 2: 4, 3: 7, 4: 10}
        start_month = quarter_starts[q]
        end_month = start_month + 2
        start = date(year, start_month, 1)
        end = date(year, end_month, calendar.monthrange(year, end_month)[1])
        label = f"Q{q} {year}"
    elif len(s) == 4 and s.isdigit():
        year = int(s)
        start = date(year, 1, 1)
        end = date(year, 12, 31)
        label = str(year)
    else:
        dt = datetime.strptime(period_str.strip(), "%Y-%m")
        start = dt.date().replace(day=1)
        end = dt.date().replace(day=calendar.monthrange(dt.year, dt.month)[1])
        label = dt.strftime("%B %Y")

    return start, end, label


# ---------------------------------------------------------------------------
# Accrual logic
# ---------------------------------------------------------------------------

def classify_accrual(row: pd.Series, period_start: date, period_end: date) -> str:
    """
    Return accrual classification:
      EARNED      – invoice date falls in period (revenue earned this period)
      ACCRUED     – no invoice yet but due date in period (accrue the revenue)
      DEFERRED    – invoiced before period but not yet due/paid (defer)
      OUT_OF_PERIOD – entirely outside the target period
    """
    inv_date = row.get("invoice_date")
    due_date = row.get("due_date")

    inv_in_period = (inv_date is not None) and (period_start <= inv_date <= period_end)
    due_in_period = (due_date is not None) and (period_start <= due_date <= period_end)

    if inv_in_period:
        return "EARNED"
    if due_in_period and (inv_date is None or inv_date > period_end):
        return "ACCRUED"
    if inv_date is not None and inv_date < period_start:
        if due_date is None or due_date >= period_start:
            return "DEFERRED"
    return "OUT_OF_PERIOD"


def calculate_accruals(df: pd.DataFrame, period_start: date, period_end: date) -> pd.DataFrame:
    """Add accrual columns to the dataframe."""
    for col in ("invoice_date", "due_date"):
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], dayfirst=True, errors="coerce").dt.date

    if "amount" not in df.columns:
        raise ValueError(
            "Could not find an amount/revenue column. "
            "Please check your file has a column like 'amount', 'revenue', or 'total'."
        )
    df["amount"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0)

    df["accrual_type"] = df.apply(classify_accrual, axis=1,
                                   period_start=period_start, period_end=period_end)
    df["accrual_amount"] = df.apply(
        lambda r: r["amount"] if r["accrual_type"] != "OUT_OF_PERIOD" else 0, axis=1
    )
    return df


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

COLOURS = {
    "header_bg":   "1F4E79",   # dark blue
    "header_fg":   "FFFFFF",
    "earned":      "C6EFCE",   # green
    "accrued":     "FFEB9C",   # amber
    "deferred":    "DAEEF3",   # light blue
    "out":         "F2F2F2",   # grey
    "total_bg":    "D9E1F2",
    "title_bg":    "2E75B6",
    "title_fg":    "FFFFFF",
    "summary_hdr": "BDD7EE",
}

TYPE_COLOURS = {
    "EARNED":        COLOURS["earned"],
    "ACCRUED":       COLOURS["accrued"],
    "DEFERRED":      COLOURS["deferred"],
    "OUT_OF_PERIOD": COLOURS["out"],
}


def _thin_border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_style(cell, bg=COLOURS["header_bg"], fg=COLOURS["header_fg"]):
    cell.font = Font(bold=True, color=fg, name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()


def _data_style(cell, bg=None, number_format=None):
    cell.font = Font(name="Arial", size=9)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(vertical="center")
    cell.border = _thin_border()
    if number_format:
        cell.number_format = number_format


def write_detail_sheet(ws, df: pd.DataFrame, period_label: str):
    # Title
    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value = f"Sales Accrual Detail — {period_label}"
    title.font = Font(bold=True, color=COLOURS["title_fg"], name="Arial", size=12)
    title.fill = PatternFill("solid", start_color=COLOURS["title_bg"])
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Determine available columns
    display_cols = []
    for col in ["invoice_id", "customer", "description", "invoice_date", "due_date",
                "amount", "status", "accrual_type", "accrual_amount"]:
        if col in df.columns or col in ("accrual_type", "accrual_amount"):
            display_cols.append(col)

    headers = {
        "invoice_id":     "Invoice / Order ID",
        "customer":       "Customer",
        "description":    "Description",
        "invoice_date":   "Invoice Date",
        "due_date":       "Due Date",
        "amount":         "Gross Amount (£)",
        "status":         "Status",
        "accrual_type":   "Accrual Type",
        "accrual_amount": "Accrual Amount (£)",
    }

    for col_idx, col in enumerate(display_cols, start=1):
        cell = ws.cell(row=2, column=col_idx, value=headers.get(col, col.replace("_", " ").title()))
        _header_style(cell)
    ws.row_dimensions[2].height = 30

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), start=3):
        bg = TYPE_COLOURS.get(row.get("accrual_type", ""), None)
        for col_idx, col in enumerate(display_cols, start=1):
            val = row.get(col, "")
            if isinstance(val, float) and pd.isna(val):
                val = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            nf = "#,##0.00" if col in ("amount", "accrual_amount") else None
            _data_style(cell, bg=bg, number_format=nf)

    # Totals row
    total_row = len(df) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name="Arial", size=9)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", start_color=COLOURS["total_bg"])
    ws.cell(row=total_row, column=1).border = _thin_border()

    for col_idx, col in enumerate(display_cols, start=1):
        cell = ws.cell(row=total_row, column=col_idx)
        if col == "amount" and col in display_cols:
            col_letter = get_column_letter(col_idx)
            cell.value = f"=SUM({col_letter}3:{col_letter}{total_row - 1})"
            cell.number_format = "#,##0.00"
        elif col == "accrual_amount":
            col_letter = get_column_letter(col_idx)
            cell.value = f"=SUM({col_letter}3:{col_letter}{total_row - 1})"
            cell.number_format = "#,##0.00"
        cell.fill = PatternFill("solid", start_color=COLOURS["total_bg"])
        cell.font = Font(bold=True, name="Arial", size=9)
        cell.border = _thin_border()

    # Column widths
    col_widths = {
        "invoice_id": 18, "customer": 22, "description": 28,
        "invoice_date": 14, "due_date": 14, "amount": 16,
        "status": 12, "accrual_type": 16, "accrual_amount": 18,
    }
    for col_idx, col in enumerate(display_cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col, 14)

    # Freeze panes
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(display_cols))}2"


def write_summary_sheet(ws, df: pd.DataFrame, period_label: str, period_start: date, period_end: date):
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    ws.merge_cells("A1:C1")
    title = ws["A1"]
    title.value = f"Accrual Summary — {period_label}"
    title.font = Font(bold=True, color=COLOURS["title_fg"], name="Arial", size=12)
    title.fill = PatternFill("solid", start_color=COLOURS["title_bg"])
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Period info
    ws["A2"] = "Period Start"
    ws["B2"] = period_start.strftime("%d %b %Y")
    ws["A3"] = "Period End"
    ws["B3"] = period_end.strftime("%d %b %Y")
    ws["A4"] = "Report Generated"
    ws["B4"] = datetime.now().strftime("%d %b %Y %H:%M")
    for r in range(2, 5):
        ws.cell(r, 1).font = Font(bold=True, name="Arial", size=9)
        ws.cell(r, 2).font = Font(name="Arial", size=9)

    # Summary table
    ws["A6"] = "Accrual Type"
    ws["B6"] = "No. of Items"
    ws["C6"] = "Total Amount (£)"
    for col in ("A6", "B6", "C6"):
        _header_style(ws[col], bg=COLOURS["summary_hdr"], fg="000000")

    accrual_types = ["EARNED", "ACCRUED", "DEFERRED", "OUT_OF_PERIOD"]
    labels = {
        "EARNED":        "Earned (invoiced in period)",
        "ACCRUED":       "Accrued (due but not yet invoiced)",
        "DEFERRED":      "Deferred (pre-invoiced, not yet due)",
        "OUT_OF_PERIOD": "Out of Period",
    }

    for row_offset, atype in enumerate(accrual_types, start=7):
        subset = df[df["accrual_type"] == atype]
        ws.cell(row_offset, 1, labels[atype]).font = Font(name="Arial", size=9)
        ws.cell(row_offset, 2, len(subset)).font = Font(name="Arial", size=9)
        ws.cell(row_offset, 3, subset["accrual_amount"].sum()).number_format = "#,##0.00"
        ws.cell(row_offset, 3).font = Font(name="Arial", size=9)
        bg = TYPE_COLOURS[atype]
        for c in (1, 2, 3):
            ws.cell(row_offset, c).fill = PatternFill("solid", start_color=bg)
            ws.cell(row_offset, c).border = _thin_border()

    # Grand total
    total_row = 7 + len(accrual_types)
    ws.cell(total_row, 1, "TOTAL ACCRUAL").font = Font(bold=True, name="Arial", size=9)
    ws.cell(total_row, 2, f"=SUM(B7:B{total_row-1})").font = Font(bold=True, name="Arial", size=9)
    total_cell = ws.cell(total_row, 3, f"=SUM(C7:C{total_row-1})")
    total_cell.number_format = "#,##0.00"
    total_cell.font = Font(bold=True, name="Arial", size=9)
    for c in (1, 2, 3):
        ws.cell(total_row, c).fill = PatternFill("solid", start_color=COLOURS["total_bg"])
        ws.cell(total_row, c).border = _thin_border()

    # Legend
    ws[f"A{total_row + 2}"] = "Legend"
    ws[f"A{total_row + 2}"].font = Font(bold=True, name="Arial", size=9)
    legend = [
        ("EARNED",        "Revenue invoiced within the target period"),
        ("ACCRUED",       "Revenue due/earned but not yet invoiced — must be accrued"),
        ("DEFERRED",      "Invoiced before period start; defer to correct period"),
        ("OUT_OF_PERIOD", "Entirely outside the target period — excluded from accrual"),
    ]
    for i, (atype, desc) in enumerate(legend, start=total_row + 3):
        ws.cell(i, 1, atype).fill = PatternFill("solid", start_color=TYPE_COLOURS[atype])
        ws.cell(i, 1).font = Font(bold=True, name="Arial", size=9)
        ws.cell(i, 1).border = _thin_border()
        ws.cell(i, 2, desc).font = Font(name="Arial", size=9)
        ws.merge_cells(f"B{i}:C{i}")


def export_to_excel(df: pd.DataFrame, output_path: str, period_label: str,
                    period_start: date, period_end: date):
    wb = Workbook()

    # Summary sheet first
    ws_summary = wb.active
    ws_summary.title = "Summary"
    write_summary_sheet(ws_summary, df, period_label, period_start, period_end)

    # Detail sheet
    ws_detail = wb.create_sheet("Accrual Detail")
    write_detail_sheet(ws_detail, df, period_label)

    # Raw data sheet
    ws_raw = wb.create_sheet("Raw Data")
    for col_idx, col in enumerate(df.columns, start=1):
        ws_raw.cell(1, col_idx, col.replace("_", " ").title()).font = Font(bold=True, name="Arial", size=9)
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col in enumerate(df.columns, start=1):
            val = row[col]
            if isinstance(val, float) and pd.isna(val):
                val = ""
            ws_raw.cell(row_idx, col_idx, val).font = Font(name="Arial", size=9)

    wb.save(output_path)
    print(f"✅  Saved: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def load_input(path: str) -> pd.DataFrame:
    p = Path(path)
    if not p.exists():
        print(f"❌  File not found: {path}")
        sys.exit(1)
    if p.suffix.lower() in (".xlsx", ".xls", ".xlsm"):
        return pd.read_excel(path, dtype=str)
    elif p.suffix.lower() == ".csv":
        return pd.read_csv(path, dtype=str)
    else:
        print(f"❌  Unsupported file type: {p.suffix}. Use .csv, .xlsx, or .xls")
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(
        description="Sales Accrual Calculator — reads invoices/orders and produces an accrual report."
    )
    parser.add_argument("--input",  "-i", required=True, help="Path to input CSV or Excel file")
    parser.add_argument("--period", "-p", required=True,
                        help="Accrual period: YYYY-MM (month), YYYY-Qn (quarter), or YYYY (year). E.g. 2025-06")
    parser.add_argument("--output", "-o", default=None,
                        help="Output Excel file path (default: accrual_<period>.xlsx)")
    args = parser.parse_args()

    try:
        period_start, period_end, period_label = parse_period(args.period)
    except ValueError as e:
        print(f"❌  Invalid period '{args.period}': {e}")
        sys.exit(1)

    output_path = args.output or f"accrual_{args.period.replace('-', '_')}.xlsx"

    print(f"📂  Loading: {args.input}")
    df = load_input(args.input)
    print(f"   {len(df)} rows loaded")

    df = normalise_columns(df)
    print(f"   Columns mapped: {list(df.columns)}")

    try:
        df = calculate_accruals(df, period_start, period_end)
    except ValueError as e:
        print(f"❌  {e}")
        sys.exit(1)

    counts = df["accrual_type"].value_counts()
    print(f"\n📊  Accrual breakdown for {period_label}:")
    for atype in ["EARNED", "ACCRUED", "DEFERRED", "OUT_OF_PERIOD"]:
        total = df.loc[df["accrual_type"] == atype, "accrual_amount"].sum()
        print(f"   {atype:<18} {counts.get(atype, 0):>4} items   £{total:>12,.2f}")

    export_to_excel(df, output_path, period_label, period_start, period_end)


if __name__ == "__main__":
    main()